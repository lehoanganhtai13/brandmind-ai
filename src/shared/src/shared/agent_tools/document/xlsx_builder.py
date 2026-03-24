"""XLSX builder for brand strategy spreadsheets using openpyxl.

Builds Excel workbooks from templates with formulas, professional
formatting, frozen panes, and brand color theming.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Color constants
_HEADER_BG = "1F4E79"
_HEADER_FG = "FFFFFF"
_BORDER_COLOR = "D9D9D9"


class BrandStrategyXLSXBuilder:
    """Builds Excel workbooks from templates with formulas.

    Supports title rows, colored headers, frozen panes,
    auto-fit column widths, and Excel formula insertion
    with placeholder resolution.
    """

    def __init__(self, brand_name: str = "Brand") -> None:
        self.brand_name = brand_name
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # type: ignore[arg-type]

    def build_from_template(
        self,
        template_config: dict[str, Any],
        data: dict[str, list[dict[str, Any]]],
    ) -> None:
        """Build workbook from template config and data.

        Args:
            template_config: Template dict from SPREADSHEET_TEMPLATES.
            data: Dict mapping sheet names to list of row dicts.
        """
        sheets = template_config.get("sheets", [])

        for sheet_config in sheets:
            sheet_name = sheet_config["name"]
            headers = sheet_config["headers"]
            formulas = sheet_config.get("formulas", {})

            ws = self.wb.create_sheet(title=sheet_name)
            sheet_data = data.get(sheet_name, [])

            self._write_title_row(ws, sheet_name, len(headers))
            self._write_headers(ws, headers)
            total_row = self._write_data_rows(ws, headers, sheet_data, formulas)
            self._apply_formatting(ws, headers, total_row)

    def _write_title_row(self, ws: Any, sheet_name: str, num_cols: int) -> None:
        """Merged row 1 with brand name + sheet title."""
        ws.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=num_cols,
        )
        cell = ws.cell(row=1, column=1)
        cell.value = f"{self.brand_name} — {sheet_name}"
        cell.font = Font(name="Calibri", size=14, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(
            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
        )

    def _write_headers(self, ws: Any, headers: list[str]) -> None:
        """Column headers with dark background and white text."""
        header_font = Font(name="Calibri", size=10, bold=True, color=_HEADER_FG)
        header_fill = PatternFill(
            start_color=_HEADER_BG,
            end_color=_HEADER_BG,
            fill_type="solid",
        )
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

    def _write_data_rows(
        self,
        ws: Any,
        headers: list[str],
        data: list[dict[str, Any]],
        formulas: dict[str, str],
    ) -> int:
        """Populate data rows and insert Excel formulas.

        Returns:
            The total row number (data end + 1).
        """
        start_row = 3
        total_row = start_row + len(data)
        last_col = get_column_letter(len(headers))
        thin_border = Border(
            left=Side(style="thin", color=_BORDER_COLOR),
            right=Side(style="thin", color=_BORDER_COLOR),
            top=Side(style="thin", color=_BORDER_COLOR),
            bottom=Side(style="thin", color=_BORDER_COLOR),
        )

        for row_idx, row_data in enumerate(data, start_row):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)

                # Check if this column has a formula
                if header in formulas:
                    formula = formulas[header]
                    resolved = (
                        formula.replace("{row}", str(row_idx))
                        .replace("{total_row}", str(total_row))
                        .replace("{row_scores}", f"B{row_idx}:{last_col}{row_idx}")
                    )
                    cell.value = resolved
                    cell.font = Font(name="Calibri", size=10)
                else:
                    cell.value = row_data.get(header, "")
                    cell.font = Font(name="Calibri", size=10)

                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        return total_row

    def _apply_formatting(self, ws: Any, headers: list[str], total_row: int) -> None:
        """Auto-fit column widths and freeze header rows."""
        ws.freeze_panes = "A3"

        for col_idx, header in enumerate(headers, 1):
            max_width = len(str(header))
            for row in ws.iter_rows(
                min_row=3,
                max_row=total_row,
                min_col=col_idx,
                max_col=col_idx,
            ):
                for cell in row:
                    if cell.value:
                        max_width = max(max_width, len(str(cell.value)))
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = min(max_width + 4, 40)

    def save(self, output_path: str) -> str:
        """Save workbook to file.

        Args:
            output_path: File path for output XLSX.

        Returns:
            Path to saved file.
        """
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(output_path)
        logger.info(f"XLSX generated: {output_path}")
        return output_path
